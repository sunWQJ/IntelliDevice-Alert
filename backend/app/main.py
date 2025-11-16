from __future__ import annotations
from fastapi import FastAPI, HTTPException
from fastapi import Body, Query, UploadFile, File
from uuid import uuid4
from .schemas import ReportIn, ReportOut, ReportStored
from .services.deidentify import remove_pii
from .services.cleaning import fingerprint
from .services.severity import classify as classify_severity, classify_with_evidence
from .storage import upsert_report, by_id, find_by_fingerprint, now
from . import audit
from .terminology import load as load_terms, search as term_search, available_categories
from typing import List, Optional
from openpyxl import load_workbook
from openpyxl import Workbook
from io import BytesIO
from . import graph
from .db import init_db
from .llm import llm_standardize, llm_structure
from fastapi.staticfiles import StaticFiles
from .services.risk_analysis import analyze_risks_in_graph
from .services.structure_analyzer import analyze_report_structure


app = FastAPI(title="IntelliDevice-Alert API", version="0.1.0")
init_db()
load_terms()
from pathlib import Path
_static = str((Path(__file__).resolve().parents[1] / "static").resolve())
app.mount("/ui", StaticFiles(directory=_static, html=True), name="static")


def _process_incoming(payload: ReportIn) -> ReportOut:
    clean = payload.dict()
    # Handle None event_datetime by providing current time as default
    if clean.get("event_datetime") is None:
        from datetime import datetime
        clean["event_datetime"] = datetime.now()
    fp = fingerprint(clean)
    dup_id = find_by_fingerprint(fp)
    report_id = str(uuid4()) if not dup_id else dup_id
    sanitized = remove_pii(clean)
    sev = sanitized.get("injury_severity") or "none"
    if not sev or sev == "none":
        lvl, ev = classify_with_evidence(sanitized.get("event_description", ""))
        sev = lvl
        try:
            import json
            audit.write(report_id, "severity_evidence", json.dumps({"level": lvl, "evidence": ev}, ensure_ascii=False))
        except Exception:
            pass
    # ensure no duplicate kw when building model
    if "injury_severity" in sanitized:
        sanitized.pop("injury_severity")
    stored = ReportStored(
        report_id=report_id,
        processed_at=now(),
        status="duplicate" if dup_id else "received",
        fingerprint=fp,
        source_version="v0",
        injury_severity=sev,
        **sanitized,
    )
    out = upsert_report(stored)
    audit.write(report_id, "received", f"status={out.status}")
    graph.write_report(out)
    return out


@app.post("/reports", response_model=ReportOut)
def create_report(
    payload: ReportIn = Body(...),
    standardize: bool = Query(False),
    categories: Optional[List[str]] = Query(None),
    top_k: int = Query(5),
    threshold: float = Query(0.3),
) -> ReportOut:
    out = _process_incoming(payload)
    if standardize:
        cats = categories or ["A", "E", "F", "G"]
        text = out.event_description
        res = term_search(text, cats, top_k=top_k, threshold=threshold)
        total = sum(len(v) for v in res.values())
        audit.write(out.report_id, "standardized", str(total))
    return out


@app.get("/reports/{report_id}", response_model=ReportOut)
def get_report(report_id: str) -> ReportOut:
    found = by_id(report_id)
    if not found:
        raise HTTPException(status_code=404, detail="Report not found")
    return found


@app.get("/reports/{report_id}/audit")
def get_report_audit(report_id: str):
    logs = audit.for_report(report_id)
    return [
        {
            "id": l.id,
            "created_at": l.created_at.isoformat(),
            "event": l.event,
            "detail": l.detail,
        }
        for l in logs
    ]


@app.post("/standardize")
def standardize(payload: dict = Body(...)):
    text = str(payload.get("text", "")).strip()
    cats = payload.get("categories") or available_categories()
    top_k = int(payload.get("top_k", 5))
    threshold = float(payload.get("threshold", 0.3))
    if not text:
        raise HTTPException(status_code=400, detail="text required")
    res = term_search(text, cats, top_k=top_k, threshold=threshold)
    out = {}
    for c, items in res.items():
        out[c] = [
            {
                "code": t.code,
                "term": t.term,
                "definition": t.definition,
                "category": t.category,
                "score": s,
            }
            for t, s in items
        ]
    return {"text": text, "results": out}


@app.post("/classify/severity")
def classify_severity_endpoint(payload: dict = Body(...)):
    text = str(payload.get("text", ""))
    if not text:
        raise HTTPException(status_code=400, detail="text required")
    lvl, ev = classify_with_evidence(text)
    return {"injury_severity": lvl, "evidence": ev}


@app.post("/analyze/failure")
def analyze_failure(payload: dict = Body(...)):
    text = str(payload.get("text", "")).strip()
    cats = payload.get("categories") or ["A", "E", "F"]
    top_k = int(payload.get("top_k", 5))
    threshold = float(payload.get("threshold", 0.3))
    if not text:
        raise HTTPException(status_code=400, detail="text required")
    res = term_search(text, cats, top_k=top_k, threshold=threshold)
    injuries = []
    for c in ("E", "F"):
        for t, s in res.get(c, []):
            injuries.append(t.term)
    lvl, ev = classify_with_evidence(text + " " + " ".join(injuries))
    out = {}
    for c, items in res.items():
        out[c] = [
            {
                "code": t.code,
                "term": t.term,
                "definition": t.definition,
                "category": t.category,
                "score": s,
            }
            for t, s in items
        ]
    return {"text": text, "severity": lvl, "evidence": ev, "matches": out}


@app.post("/reports/structured")
def create_structured_report(payload: dict = Body(...)):
    init_db()
    base = {
        "hospital_id": payload.get("hospital_id"),
        "device_name": payload.get("device_name"),
        "manufacturer": payload.get("manufacturer"),
        "model": payload.get("model"),
        "lot_sn": payload.get("lot_sn"),
        "event_datetime": payload.get("event_datetime"),
        "event_description": payload.get("event_description", ""),
        "action_taken": payload.get("action_taken"),
    }
    failure_desc = str(payload.get("failure_desc", ""))
    injury_desc = str(payload.get("injury_desc", ""))
    action_desc = str(payload.get("action_desc", ""))

    out = _process_incoming(ReportIn(**base))
    cats_map = {
        "FailureMode": ["A"],
        "Injury": ["E", "F"],
        "Action": ["C", "D"],
    }
    matches = {}
    if failure_desc:
        res = term_search(failure_desc, cats_map["FailureMode"], top_k=1, threshold=0.3)
        m = None
        for c in cats_map["FailureMode"]:
            if res.get(c):
                m = res[c][0]
                break
        if m:
            t, s = m
            matches["FailureMode"] = {"code": t.code, "term": t.term, "definition": t.definition, "category": t.category}
    if injury_desc:
        res = term_search(injury_desc, cats_map["Injury"], top_k=1, threshold=0.3)
        m = None
        for c in cats_map["Injury"]:
            if res.get(c):
                m = res[c][0]
                break
        if m:
            t, s = m
            matches["Injury"] = {"code": t.code, "term": t.term, "definition": t.definition, "category": t.category}
            lvl, ev = classify_with_evidence(injury_desc)
            try:
                import json
                audit.write(out.report_id, "severity_evidence", json.dumps({"level": lvl, "evidence": ev}, ensure_ascii=False))
            except Exception:
                pass
    if action_desc:
        res = term_search(action_desc, cats_map["Action"], top_k=1, threshold=0.3)
        m = None
        for c in cats_map["Action"]:
            if res.get(c):
                m = res[c][0]
                break
        if m:
            t, s = m
            matches["Action"] = {"code": t.code, "term": t.term, "definition": t.definition, "category": t.category}

    # store entities in DB (in-memory)
    try:
        from .db import get_db
        db = get_db()
        # Store entities in the report data
        report_data = db.get_report(out.report_id)
        if report_data:
            if 'entities' not in report_data:
                report_data['entities'] = []
            for et, m in matches.items():
                report_data['entities'].append({
                    'entity_type': et, 
                    'code': m.get("code"), 
                    'term': m.get("term"), 
                    'definition': m.get("definition"), 
                    'category': m.get("category")
                })
            db.add_report(out.report_id, report_data)
    except Exception as e:
        import logging
        logging.error(f"Failed to store entities: {e}")

    # write to graph
    structure = {"entities": [], "relations": []}
    for et, m in matches.items():
        structure["entities"].append({"type": et, "code": m.get("code"), "term": m.get("term")})
    if matches.get("FailureMode") and matches.get("Injury"):
        structure["relations"].append({"type": "CAUSES", "from": matches["FailureMode"]["term"], "to": matches["Injury"]["term"]})
    graph.write_structure(out.report_id, structure)
    return {"report": out, "matches": matches}


@app.get("/reports/{report_id}/standardized")
def standardized_for_report(report_id: str, categories: Optional[List[str]] = None, top_k: int = 5, threshold: float = 0.3):
    found = by_id(report_id)
    if not found:
        raise HTTPException(status_code=404, detail="Report not found")
    cats = categories or ["A", "E", "F", "G"]
    text = found.event_description
    res = term_search(text, cats, top_k=top_k, threshold=threshold)
    out = {}
    for c, items in res.items():
        out[c] = [
            {
                "code": t.code,
                "term": t.term,
                "definition": t.definition,
                "category": t.category,
                "score": s,
            }
            for t, s in items
        ]
    return {"report_id": report_id, "text": text, "results": out}


@app.post("/llm/standardize")
def llm_standardize_endpoint(payload: dict = Body(...)):
    provider = str(payload.get("provider", "openai"))
    text = str(payload.get("text", ""))
    categories = payload.get("categories")
    top_k = int(payload.get("top_k", 5))
    if not text:
        raise HTTPException(status_code=400, detail="text required")
    data = llm_standardize(provider, text, categories, top_k)
    if "error" in data:
        raise HTTPException(status_code=503, detail=data["error"])
    return data


@app.post("/llm/structure")
def llm_structure_endpoint(payload: dict = Body(...)):
    provider = str(payload.get("provider", "openai"))
    text = str(payload.get("text", ""))
    top_k = int(payload.get("top_k", 5))
    if not text:
        raise HTTPException(status_code=400, detail="text required")
    data = llm_structure(provider, text, top_k)
    if "error" in data:
        raise HTTPException(status_code=503, detail=data["error"])
    return data


@app.post("/llm/structure/store")
def llm_structure_store(payload: dict = Body(...)):
    provider = str(payload.get("provider", "openai"))
    text = str(payload.get("text", ""))
    report_id = str(payload.get("report_id", ""))
    top_k = int(payload.get("top_k", 5))
    if not text or not report_id:
        raise HTTPException(status_code=400, detail="text and report_id required")
    data = llm_structure(provider, text, top_k)
    if "error" in data:
        raise HTTPException(status_code=503, detail=data["error"])
    out = graph.write_structure(report_id, data)
    if "error" in out:
        raise HTTPException(status_code=503, detail="neo4j_unavailable")
    return {"stored": out, "structure": data}


@app.post("/config/llm")
def set_llm_config(payload: dict = Body(...)):
    import os
    provider = str(payload.get("provider", "")).lower()
    api_key = str(payload.get("api_key", ""))
    model = str(payload.get("model", ""))
    if provider == "openai":
        if api_key:
            os.environ["OPENAI_API_KEY"] = api_key
        if model:
            os.environ["OPENAI_MODEL"] = model
        return {"ok": True, "provider": "openai"}
    elif provider == "gemini":
        if api_key:
            os.environ["GEMINI_API_KEY"] = api_key
        if model:
            os.environ["GEMINI_MODEL"] = model
        return {"ok": True, "provider": "gemini"}
    else:
        raise HTTPException(status_code=400, detail="unsupported provider")


@app.post("/config/neo4j")
def set_neo4j_config(payload: dict = Body(...)):
    import os
    uri = str(payload.get("uri", ""))
    user = str(payload.get("user", ""))
    password = str(payload.get("password", ""))
    if uri:
        os.environ["NEO4J_URI"] = uri
    if user:
        os.environ["NEO4J_USER"] = user
    if password:
        os.environ["NEO4J_PASS"] = password
    return {"ok": True}


@app.get("/dashboard/summary")
def dashboard_summary():
    from .db import get_db
    import collections
    
    db = get_db()
    reports = db.get_all_reports()
    
    total = len(reports)
    top_devices = collections.Counter([r.get('device_name') for r in reports]).most_common(5)
    top_severity = collections.Counter([r.get('injury_severity') for r in reports]).most_common(5)
    recent = [
        {
            "report_id": r.get('report_id'),
            "hospital_id": r.get('hospital_id'),
            "device_name": r.get('device_name'),
            "injury_severity": r.get('injury_severity'),
            "event_datetime": r.get('event_datetime'),
        }
        for r in reports[:10]
    ]
    return {"total": total, "top_devices": top_devices, "top_severity": top_severity, "recent": recent}


@app.get("/case/{report_id}/graph")
def case_graph(report_id: str):
    print(f"Looking for report_id: {report_id}")
    r = by_id(report_id)
    if not r:
        print(f"Report not found: {report_id}")
        raise HTTPException(status_code=404, detail="Report not found")
    print(f"Found report: {r.report_id}")
    # 尝试从 Neo4j 读取扩展关系
    try:
        g = graph.case_graph(report_id)
    except Exception as e:
        print(f"Neo4j graph query failed: {e}")
        g = {"error": str(e)}
        if "error" in g:
            # 回退到DB构建视图（包含故障/伤害/处置、制造商/型号/事件日期）
            nodes = {
                r.report_id: {"id": r.report_id, "label": "Report", "name": r.report_id},
                r.hospital_id: {"id": r.hospital_id, "label": "Hospital", "name": r.hospital_id},
                r.device_name: {"id": r.device_name, "label": "Device", "name": r.device_name},
            }
            edges = set()
            edges.add((r.report_id, r.hospital_id, "REPORTED_BY"))
            edges.add((r.report_id, r.device_name, "RESULTS_IN"))
            if r.manufacturer:
                nodes[r.manufacturer] = {"id": r.manufacturer, "label": "Manufacturer", "name": r.manufacturer}
                edges.add((r.device_name, r.manufacturer, "MANUFACTURED_BY"))
            if r.model:
                nodes[r.model] = {"id": r.model, "label": "Model", "name": r.model}
                edges.add((r.device_name, r.model, "HAS_MODEL"))
            if r.event_datetime:
                dt = r.event_datetime.isoformat()
                nodes[dt] = {"id": dt, "label": "EventDate", "name": dt}
                edges.add((r.report_id, dt, "AT_TIME"))
            # 处置措施节点
            if getattr(r, 'action_taken', None):
                act = r.action_taken
                nodes[act] = {"id": act, "label": "Action", "name": act}
                edges.add((r.report_id, act, "HAS_ACTION"))
        # 从 report_entities 读取术语映射（in-memory）
        try:
            from .db import get_db
            db = get_db()
            report_data = db.get_report(report_id)
            if report_data and 'entities' in report_data:
                for entity in report_data['entities']:
                    nid = entity.get('term')
                    et = entity.get('entity_type')
                    if not nid or not et:
                        continue
                    # 统一实体标签用于图谱展示
                    et_map = {
                        'FAILURE_MODE': 'FailureMode',
                        'FailureMode': 'FailureMode',
                        'HEALTH_IMPACT': 'Injury',
                        'HARM': 'Injury',
                        'Injury': 'Injury',
                        'CLINICAL_MANIFESTATION': 'ClinicalManifestation',
                        'SYMPTOM': 'ClinicalManifestation',
                        'DeviceIssue': 'DeviceIssue'
                    }
                    label = et_map.get(et, et)
                    nodes[nid] = {
                        "id": nid,
                        "label": label,
                        "name": entity.get('term'),
                        "code": entity.get('code'),
                        "definition": entity.get('definition'),
                        "category": entity.get('category')
                    }
                    edges.add((r.report_id, nid, f"HAS_{label.upper()}"))
        except Exception as e:
            import logging
            logging.error(f"Failed to load entities from in-memory DB: {e}")
                # 若存在故障与伤害，补充CAUSES关系
        if report_data and 'entities' in report_data:
            fm = [e for e in report_data['entities'] if e.get('entity_type') in ("FailureMode", "FAILURE_MODE")] 
            inj = [e for e in report_data['entities'] if e.get('entity_type') in ("Injury", "HEALTH_IMPACT", "HARM")]
            if fm and inj:
                edges.add((fm[0]['term'], inj[0]['term'], "CAUSES"))
        # 若有结构化摘要，补充节点与关系
        if report_data and 'structured_data' in report_data:
            sd = report_data['structured_data'] or {}
            if sd.get('injury'):
                inj_name = sd['injury']
                nodes[inj_name] = {"id": inj_name, "label": "Injury", "name": inj_name}
                edges.add((r.report_id, inj_name, "HAS_INJURY"))
            if sd.get('failure'):
                fm_name = sd['failure']
                nodes[fm_name] = {"id": fm_name, "label": "FailureMode", "name": fm_name}
                edges.add((r.device_name, fm_name, "HAS_FAILUREMODE"))
                if sd.get('injury'):
                    edges.add((fm_name, sd['injury'], "CAUSES"))
            if sd.get('device_issue'):
                di_name = sd['device_issue']
                nodes[di_name] = {"id": di_name, "label": "DeviceIssue", "name": di_name}
                edges.add((r.device_name, di_name, "HAS_FAULT"))
        return {"nodes": list(nodes.values()), "edges": [{"source": s, "target": t, "label": l} for (s, t, l) in edges]}
    return g


@app.get("/case/recent-graph")
def case_recent_graph(limit: int = 10):
    """获取最近案例的知识图谱，包含重要的医疗信息如伤害、故障等"""
    from .db import get_db
    import logging
    
    try:
        db = get_db()
        reports = db.get_reports_by_limit(limit)
        
        nodes_map = {}
        edges_set = set()
        
        for r in reports:
            try:
                report_id = r.get('report_id')
                if not report_id:
                    continue
                    
                # Add basic report information
                nodes_map[report_id] = {
                    "id": report_id, 
                    "label": "AdverseEventReport", 
                    "name": report_id, 
                    "severity": r.get('injury_severity', 'Unknown')
                }
                
                # Add hospital information
                hospital_id = r.get('hospital_id', 'Unknown_Hospital')
                nodes_map[hospital_id] = {
                    "id": hospital_id, 
                    "label": "Hospital", 
                    "name": hospital_id
                }
                edges_set.add((report_id, hospital_id, "REPORTED_BY"))
                
                # Add device information
                device_name = r.get('device_name', 'Unknown_Device')
                nodes_map[device_name] = {
                    "id": device_name, 
                    "label": "MedicalDevice", 
                    "name": device_name
                }
                edges_set.add((device_name, report_id, "RELATED_TO"))
                
                # Add manufacturer information
                if r.get('manufacturer'):
                    manufacturer = r.get('manufacturer')
                    nodes_map[manufacturer] = {
                        "id": manufacturer, 
                        "label": "Manufacturer", 
                        "name": manufacturer
                    }
                    edges_set.add((device_name, manufacturer, "MANUFACTURED_BY"))
                    
                # Add model information
                if r.get('model'):
                    model = r.get('model')
                    nodes_map[model] = {
                        "id": model, 
                        "label": "Model", 
                        "name": model
                    }
                    edges_set.add((device_name, model, "HAS_MODEL"))
                
                # Add date/time information
                if r.get('event_datetime'):
                    dt = str(r.get('event_datetime'))
                    nodes_map[dt] = {
                        "id": dt, 
                        "label": "DiscoveryDate", 
                        "name": dt
                    }
                    edges_set.add((report_id, dt, "AT_TIME"))
                
                # Add important medical information from entities (伤害、故障等)
                entities = r.get('entities', [])
                print(f"Report {report_id} has {len(entities)} entities")
                for entity in entities:
                    entity_type = entity.get('entity_type')
                    term = entity.get('term')
                    category = entity.get('category')
                    
                    if term and entity_type:
                        # Create node for the medical concept
                        node_id = f"{entity_type}_{term}"
                        nodes_map[node_id] = {
                            "id": node_id,
                            "label": entity_type,
                            "name": term,
                            "category": category,
                            "definition": entity.get('definition', '')
                        }
                        
                        # Connect to the report based on entity type
                        if entity_type in ['HEALTH_IMPACT', 'INJURY', 'HARM']:
                            edges_set.add((report_id, node_id, "CAUSED"))
                        elif entity_type in ['DEVICE_ISSUE', 'FAILURE_MODE', 'MALFUNCTION']:
                            edges_set.add((device_name, node_id, "EXPERIENCED"))
                        elif entity_type in ['CLINICAL_MANIFESTATION', 'SYMPTOM']:
                            edges_set.add((report_id, node_id, "PRESENTED_WITH"))
                        else:
                            edges_set.add((report_id, node_id, "CONTAINS"))
                
                # Add structured data if available
                if r.get('structured_data'):
                    print(f"Report {report_id} has structured data: {r.get('structured_data')}")
                    structured = r.get('structured_data')
                    
                    # Add injury information (伤害)
                    if structured.get('injury'):
                        injury = structured.get('injury')
                        injury_id = f"INJURY_{injury}"
                        nodes_map[injury_id] = {
                            "id": injury_id,
                            "label": "Injury",
                            "name": injury
                        }
                        edges_set.add((report_id, injury_id, "RESULTED_IN"))
                    
                    # Add failure information (故障)
                    if structured.get('failure'):
                        failure = structured.get('failure')
                        failure_id = f"FAILURE_{failure}"
                        nodes_map[failure_id] = {
                            "id": failure_id,
                            "label": "Failure",
                            "name": failure
                        }
                        edges_set.add((device_name, failure_id, "EXPERIENCED"))
                    
                    # Add device issue information
                    if structured.get('device_issue'):
                        issue = structured.get('device_issue')
                        issue_id = f"ISSUE_{issue}"
                        nodes_map[issue_id] = {
                            "id": issue_id,
                            "label": "DeviceIssue",
                            "name": issue
                        }
                        edges_set.add((device_name, issue_id, "HAS_ISSUE"))
                else:
                    print(f"Report {report_id} has no structured data")
                        
            except Exception as e:
                logging.error(f"Error processing report {r.get('report_id', 'unknown')}: {e}")
                continue
        
        nodes = list(nodes_map.values())
        edges = [{"source": s, "target": t, "label": l} for (s, t, l) in edges_set]
        
        logging.info(f"Generated overview graph with {len(nodes)} nodes and {len(edges)} edges")
        
        return {"nodes": nodes, "edges": edges}
    except Exception as e:
        import logging
        logging.error(f"Error in case_recent_graph: {e}")
        return {"nodes": [], "edges": []}


@app.post("/reports/analyze-structure")
def analyze_report_structure_endpoint(payload: dict = Body(...), use_llm: bool = Query(False)):
    """分析报告结构并提取关键信息"""
    try:
        # 验证必要字段
        required_fields = ["event_description"]
        for field in required_fields:
            if not payload.get(field):
                return {
                    "success": False,
                    "error": f"缺少必要字段：{field}"
                }
        
        result = analyze_report_structure(payload)
        if use_llm:
            provider = payload.get("provider", "openai")
            text = payload.get("event_description", "")
            llm_out = llm_structure(provider=provider, text=text, top_k=5)
            result["llm_entities"] = llm_out.get("entities", []) if isinstance(llm_out, dict) else []
            result["llm_relations"] = llm_out.get("relations", []) if isinstance(llm_out, dict) else []
            result["llm_error"] = llm_out.get("error") if isinstance(llm_out, dict) else None
        return {"success": True, "data": result}
    except Exception as e:
        return {
            "success": False,
            "error": f"分析失败：{str(e)}"
        }


@app.post("/reports/structured-confirm")
def confirm_structured_report(payload: dict = Body(...)):
    """确认结构化数据并录入系统"""
    try:
        # 首先进行结构化分析
        structure_result = analyze_report_structure(payload)
        
        # 创建标准报告
        base_report = {
            "hospital_id": payload.get("hospital_id"),
            "device_name": payload.get("device_name"),
            "manufacturer": payload.get("manufacturer"),
            "model": payload.get("model"),
            "lot_sn": payload.get("lot_sn"),
            "event_datetime": payload.get("event_datetime"),
            "event_description": payload.get("event_description"),
            "injury_severity": payload.get("injury_severity", "none"),
            "action_taken": payload.get("action_taken"),
        }
        
        # 处理报告
        report_out = _process_incoming(ReportIn(**base_report))
        
        # 存储结构化实体（in-memory）
        try:
            from .db import get_db
            db = get_db()
            report_data = db.get_report(report_out.report_id)
            if report_data:
                if 'entities' not in report_data:
                    report_data['entities'] = []
                
                # 存储匹配的标准术语（标准化实体类型）
                def _map_entity_type(field: str) -> str:
                    f = str(field).lower()
                    if f == 'failure_mode':
                        return 'FailureMode'
                    if f == 'health_impact':
                        return 'Injury'
                    if f == 'clinical_manifestation':
                        return 'ClinicalManifestation'
                    if f == 'device_issue':
                        return 'DeviceIssue'
                    return field.upper()
                for term_info in structure_result.get("matched_terms", []):
                    report_data['entities'].append({
                        'entity_type': _map_entity_type(term_info.get("field")),
                        'code': term_info.get("code"),
                        'term': term_info.get("term"),
                        'definition': "",
                        'category': term_info.get("category")
                    })
                # 若无标准术语匹配，按抽取结果补充基础实体
                def _add_basic(et: str, name: str):
                    if name and name not in ("未知故障模式", "临床表现未明确", "健康影响未明确", "处置措施未明确"):
                        report_data['entities'].append({
                            'entity_type': et,
                            'code': None,
                            'term': name,
                            'definition': "",
                            'category': None
                        })
                _add_basic('FailureMode', structure_result.get('failure_mode'))
                _add_basic('Injury', structure_result.get('health_impact'))
                _add_basic('ClinicalManifestation', structure_result.get('clinical_manifestation'))
                _add_basic('DeviceIssue', structure_result.get('device_issue'))
                
                # 存储结构化分析结果
                report_data['entities'].append({
                    'entity_type': "STRUCTURE_ANALYSIS",
                    'code': "ANALYSIS",
                    'term': "结构化分析结果",
                    'definition': str(structure_result),
                    'category': "ANALYSIS"
                })
                
                # 保存结构化摘要，供Overview Graph消费
                report_data['structured_data'] = {
                    'device_issue': structure_result.get('device_issue'),
                    'failure': structure_result.get('failure_mode'),
                    'injury': structure_result.get('health_impact'),
                    'clinical_manifestation': structure_result.get('clinical_manifestation')
                }
                db.add_report(report_out.report_id, report_data)
        except Exception as e:
            print(f"存储结构化实体失败：{e}")
        
        structure = {"entities": [], "relations": []}
        llm_entities = payload.get("llm_entities")
        llm_relations = payload.get("llm_relations")
        if isinstance(llm_entities, list) or isinstance(llm_relations, list):
            structure["entities"] = llm_entities or []
            structure["relations"] = llm_relations or []
        else:
            for term_info in structure_result.get("matched_terms", []):
                structure["entities"].append({
                    "type": term_info["field"].upper(),
                    "code": term_info["code"],
                    "term": term_info["term"]
                })
            failure_modes = [t for t in structure_result.get("matched_terms", []) if t["field"] == "failure_mode"]
            injuries = [t for t in structure_result.get("matched_terms", []) if t["field"] == "health_impact"]
            if failure_modes and injuries:
                structure["relations"].append({
                    "type": "CAUSES",
                    "from": failure_modes[0]["term"],
                    "to": injuries[0]["term"]
                })
            # 基础抽取兜底：即使未命中术语库也写入故障/伤害/设备问题
            base_failure = structure_result.get("failure_mode")
            base_injury = structure_result.get("health_impact")
            base_issue = structure_result.get("device_issue")
            if base_failure:
                structure["entities"].append({
                    "type": "FAILURE_MODE",
                    "code": None,
                    "term": base_failure
                })
            if base_injury:
                structure["entities"].append({
                    "type": "INJURY",
                    "code": None,
                    "term": base_injury
                })
            if base_issue:
                structure["entities"].append({
                    "type": "DEVICE_ISSUE",
                    "code": None,
                    "term": base_issue
                })
            if base_failure and base_injury:
                structure["relations"].append({
                    "type": "CAUSES",
                    "from": base_failure,
                    "to": base_injury
                })

        # 写入图数据库
        graph.write_structure(report_out.report_id, structure)
        
        # 记录审计日志
        import json
        audit.write(report_out.report_id, "structured_confirmed", json.dumps({
            "confidence": structure_result.get("analysis_confidence", 0),
            "matched_terms_count": len(structure_result.get("matched_terms", [])),
            "device_issue": structure_result.get("device_issue"),
            "failure_mode": structure_result.get("failure_mode"),
            "clinical_manifestation": structure_result.get("clinical_manifestation"),
            "health_impact": structure_result.get("health_impact"),
            "treatment_action": structure_result.get("treatment_action")
        }, ensure_ascii=False))
        
        return {
            "success": True,
            "data": {
                "report": report_out,
                "structure": structure_result,
                "entities_stored": len(structure.get("entities", [])),
                "confidence": structure_result.get("analysis_confidence", 0)
            }
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": f"确认录入失败：{str(e)}"
        }


@app.post("/llm/restructure")
def restructure_with_llm(payload: dict = Body(...)):
    """使用LLM重新构建和优化报告内容"""
    try:
        provider = payload.get("provider", "openai")
        
        # 构建提示词
        prompt = f"""
        请根据以下医疗事件报告信息，重新组织和优化描述，使其更清晰、准确：
        
        设备名称：{payload.get('device_name', '')}
        事件描述：{payload.get('event_description', '')}
        处理措施：{payload.get('action_taken', '')}
        
        要求：
        1. 保持医疗专业性和准确性
        2. 突出关键信息（设备问题、患者影响、处置措施）
        3. 使用标准医疗术语
        4. 语言简洁明了
        
        请返回优化后的JSON格式数据，包含：device_name, event_description, action_taken
        """
        
        # 调用LLM服务
        from .llm import llm_standardize
        
        llm_result = llm_standardize(
            text=prompt,
            provider=provider,
            categories=[],
            top_k=1
        )
        
        # 解析LLM响应
        if llm_result and "results" in llm_result:
            # 提取优化后的内容
            optimized_content = llm_result["results"]
            
            return {
                "success": True,
                "data": {
                    "device_name": payload.get("device_name"),  # 保持原设备名称
                    "event_description": optimized_content.get("event_description", payload.get("event_description")),
                    "action_taken": optimized_content.get("action_taken", payload.get("action_taken"))
                }
            }
        else:
            return {
                "success": False,
                "error": "LLM处理失败"
            }
            
    except Exception as e:
        return {
            "success": False,
            "error": f"LLM重构失败：{str(e)}"
        }


@app.post("/graph/risk-analysis")
def graph_risk_analysis(payload: dict = Body(...)):
    """分析图数据中的风险点"""
    try:
        # 获取图数据
        limit = payload.get("limit", 50)
        graph_data = case_recent_graph(limit=limit)
        
        # 分析风险点
        risk_results = analyze_risks_in_graph(graph_data)
        
        return {
            "success": True,
            "data": risk_results,
            "graph_summary": {
                "total_nodes": len(graph_data.get("nodes", [])),
                "total_edges": len(graph_data.get("edges", [])),
                "analysis_limit": limit
            }
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "data": {
                "total_risks": 0,
                "high_risks": 0,
                "medium_risks": 0,
                "low_risks": 0,
                "risk_details": []
            }
        }


@app.get("/evaluate/terms")
def evaluate_terms(file_path: str = "IMDRF测试集.json", category: str = "E", top_k: int = 5, threshold: float = 0.0):
    import json
    from pathlib import Path
    p = Path(file_path)
    if not p.exists():
        raise HTTPException(status_code=404, detail="dataset not found")
    with p.open("r", encoding="utf-8") as fp:
        data = json.load(fp)
    samples = [
        {"text": x.get("伤害描述", ""), "code": x.get("正确IMDRF编码", ""), "term": x.get("术语", "")}
        for x in data
    ]
    k_list = [1, 3, top_k]
    correct = {k: 0 for k in k_list}
    mrr_sum = 0.0
    mismatches = []
    for s in samples:
        res = term_search(s["text"], [category], top_k=top_k, threshold=threshold).get(category, [])
        codes = [t.code for (t, sc) in res]
        terms = [t.term for (t, sc) in res]
        gold_code = s["code"].strip()
        gold_term = s["term"].strip()
        rank = None
        for i, (c, tm) in enumerate(zip(codes, terms), start=1):
            if c == gold_code or tm == gold_term:
                rank = i
                break
        for k in k_list:
            if rank is not None and rank <= k:
                correct[k] += 1
        if rank is not None:
            mrr_sum += 1.0 / rank
        else:
            mismatches.append({
                "text": s["text"],
                "gold_code": gold_code,
                "gold_term": gold_term,
                "pred_top1": {"code": codes[0] if codes else None, "term": terms[0] if terms else None}
            })
    total = len(samples)
    return {
        "samples": total,
        "hit@1": correct[1] / total,
        "hit@3": correct[3] / total,
        "hit@k": correct[top_k] / total,
        "mrr": mrr_sum / total,
        "mismatches": mismatches[:50]
    }


@app.post("/graph/import-terms")
def graph_import_terms(payload: dict = Body(...)):
    import logging
    from .config import terminology_dir
    
    logging.info(f"术语导入请求 - payload: {payload}")
    
    base = terminology_dir()
    if not base:
        logging.error("术语库目录未找到")
        raise HTTPException(status_code=404, detail="terminology dir not found")
    
    logging.info(f"术语库目录: {base}")
    
    cats = payload.get("categories")
    include_syn = bool(payload.get("include_synonyms", True))
    
    logging.info(f"导入参数 - categories: {cats}, include_synonyms: {include_syn}")
    
    # 检查Neo4j连接
    from .graph import _get_driver
    driver = _get_driver()
    if driver is None:
        logging.error("Neo4j驱动为None，无法连接数据库")
        raise HTTPException(status_code=503, detail="neo4j_unavailable")
    
    logging.info("Neo4j驱动正常，开始导入术语")
    
    out = graph.import_standard_terms(base, categories=cats, include_synonyms=include_syn)
    if "error" in out:
        logging.error(f"术语导入失败: {out}")
        raise HTTPException(status_code=503, detail="neo4j_unavailable")
    
    logging.info(f"术语导入成功: {out}")
    return out


@app.get("/graph/term/{code}/neighbors")
def graph_term_neighbors(code: str):
    out = graph.neighbors_by_code(code)
    if "error" in out:
        raise HTTPException(status_code=503, detail="neo4j_unavailable")
    return out


@app.get("/graph/status")
def graph_status():
    return graph.status()

@app.get("/reports/review-pending")
def review_pending():
    from .db import get_db
    items = get_db().list_pending()
    return {"items": items}

@app.post("/reports/review-confirm")
def review_confirm(payload: dict = Body(...)):
    rid = str(payload.get("report_id", ""))
    use_llm = bool(payload.get("use_llm", False))
    provider = str(payload.get("provider", "openai"))
    from .db import get_db
    item = get_db().pop_pending(rid)
    if not item:
        return {"success": False, "error": "not_found"}
    base = item.get("base") or {}
    sr = item.get("structure") or {}
    p = ReportIn(**base)
    out = _process_incoming(p)
    if use_llm:
        txt = str(p.event_description or "")
        llm_out = llm_structure(provider=provider, text=txt, top_k=5)
        ents = llm_out.get("entities", []) if isinstance(llm_out, dict) else []
        rels = llm_out.get("relations", []) if isinstance(llm_out, dict) else []
        graph.write_structure(out.report_id, {"entities": ents, "relations": rels})
    else:
        structure = {"entities": [], "relations": []}
        for term_info in sr.get("matched_terms", []):
            structure["entities"].append({"type": term_info["field"].upper(), "code": term_info["code"], "term": term_info["term"]})
        fms = [t for t in sr.get("matched_terms", []) if t["field"] == "failure_mode"]
        inj = [t for t in sr.get("matched_terms", []) if t["field"] == "health_impact"]
        if fms and inj:
            structure["relations"] = [{"type": "CAUSES", "from": fms[0]["term"], "to": inj[0]["term"]}]
        if sr.get("failure_mode"):
            structure["entities"].append({"type": "FAILURE_MODE", "code": None, "term": sr.get("failure_mode")})
        if sr.get("health_impact"):
            structure["entities"].append({"type": "INJURY", "code": None, "term": sr.get("health_impact")})
        if sr.get("device_issue"):
            structure["entities"].append({"type": "DEVICE_ISSUE", "code": None, "term": sr.get("device_issue")})
        graph.write_structure(out.report_id, structure)
    return {"success": True, "report_id": out.report_id}
@app.post("/reports/upload-excel")
async def upload_excel(file: UploadFile = File(...), auto_threshold: float = Query(0.6), review_threshold: float = Query(0.3), review_queue: bool = Query(True)):
    try:
        content = await file.read()
        wb = load_workbook(filename=BytesIO(content), read_only=True)
        ws = wb.active
    except Exception as e:
        return {"summary": {"received": 0, "duplicate": 0, "failed": 1}, "error": f"excel_read_error: {str(e)}"}
    headers_raw = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    def norm(s: str) -> str:
        return str(s).strip()
    zh_map = {
        "医院名称": "hospital_id",
        "设备名称": "device_name",
        "制造商": "manufacturer",
        "型号": "model",
        "批次或序列号": "lot_sn",
        "事件时间": "event_datetime",
        "事件描述": "event_description",
        "伤害严重度": "injury_severity",
        "处置措施": "action_taken",
    }
    headers = [zh_map.get(norm(h), h) for h in headers_raw]
    results = {"received": 0, "duplicate": 0, "failed": 0, "pending": 0, "auto_confirmed": 0}
    ids: List[str] = []
    pending_ids: List[str] = []
    for row in ws.iter_rows(min_row=2):
        data = {}
        for i, cell in enumerate(row):
            key = headers[i] if i < len(headers) else ""
            if not key:
                continue
            val = cell.value
            # Handle Excel datetime conversion
            if key == "event_datetime" and val is not None:
                if isinstance(val, (int, float)):
                    # Excel serial date number
                    from datetime import datetime, timedelta
                    val = datetime(1899, 12, 30) + timedelta(days=val)
                elif isinstance(val, str):
                    # String date - try to parse
                    try:
                        from datetime import datetime
                        val = datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        try:
                            val = datetime.strptime(val, "%Y-%m-%d")
                        except ValueError:
                            try:
                                val = datetime.strptime(val, "%Y/%m/%d %H:%M:%S")
                            except ValueError:
                                try:
                                    val = datetime.strptime(val, "%Y/%m/%d")
                                except ValueError:
                                    val = None  # Will be handled by default in ReportIn
            if key == "injury_severity" and isinstance(val, str):
                s = val.strip().lower()
                cn_map = {
                    "轻微": "mild",
                    "轻度": "mild",
                    "中度": "moderate",
                    "中等": "moderate",
                    "重度": "severe",
                    "严重": "severe",
                    "死亡": "death",
                    "无伤害": "none",
                    "无": "none"
                }
                val = cn_map.get(s, s)
            if key == "event_description" and (val is None or isinstance(val, (int, float))):
                val = str(val) if val is not None else ""
            data[key] = val
        try:
            # Validate required fields
            if not data.get("hospital_id") or not data.get("device_name") or not data.get("event_description"):
                print(f"Missing required fields in row: {data}")
                results["failed"] += 1
                continue
            
            payload = ReportIn(**data)
            out = _process_incoming(payload)
            results[out.status] += 1
            ids.append(out.report_id)
            sr = analyze_report_structure(payload.dict())
            score = float(sr.get("analysis_confidence", 0) or 0)
            if score >= auto_threshold:
                confirm_structured_report({**payload.dict(), "llm_entities": None, "llm_relations": None})
                results["auto_confirmed"] += 1
            else:
                if review_queue:
                    from .db import get_db
                    suggest_llm = score < review_threshold
                    get_db().add_pending(out.report_id, {"base": payload.dict(), "structure": sr, "score": score, "suggest_llm": suggest_llm})
                    results["pending"] += 1
                    pending_ids.append(out.report_id)
        except Exception as e:
            results["failed"] += 1
            print(f"Excel row failed: {str(e)}")
            print(f"Problematic data: {data}")
            import traceback
            traceback.print_exc()
    return {"summary": results, "report_ids": ids, "pending_ids": pending_ids}


@app.get("/templates/reports.xlsx")
def reports_template(lang: str = "zh"):
    wb = Workbook()
    ws = wb.active
    if lang == "zh":
        headers = [
            "医院编号",
            "器械名称",
            "制造商",
            "型号",
            "批次或序列号",
            "事件时间",
            "事件描述",
            "伤害严重度",
            "处置措施",
        ]
        ws.append(headers)
        ws.append([
            "H001",
            "心电监护仪",
            "ACME",
            "ECG-200",
            "L123",
            "2025-01-01T10:20:30Z",
            "设备屏幕无显示，患者监护中断",
            "moderate",
            "更换设备并加护观察",
        ])
    else:
        headers = [
            "hospital_id",
            "device_name",
            "manufacturer",
            "model",
            "lot_sn",
            "event_datetime",
            "event_description",
            "injury_severity",
            "action_taken",
        ]
        ws.append(headers)
        ws.append([
            "H001",
            "ECG Monitor",
            "ACME",
            "ECG-200",
            "L123",
            "2025-01-01T10:20:30Z",
            "Screen blank during monitoring",
            "moderate",
            "Replace device and ICU observation",
        ])
    import tempfile, os
    from fastapi.responses import FileResponse
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    fname = "MDAE模板.xlsx" if lang == "zh" else "MDAE_template.xlsx"
    return FileResponse(path=tmp.name, filename=fname, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
@app.post("/admin/db/clear")
def admin_db_clear():
    from .db import get_db
    db = get_db()
    before = db.clear()
    return {"ok": True, "cleared": before}

@app.post("/reports/import-json")
def import_json(payload: dict = Body(None), path: Optional[str] = Query(None), auto_threshold: float = Query(0.6), review_queue: bool = Query(False)):
    from pathlib import Path
    import json as _json
    base_path = Path(path or Path(__file__).resolve().parents[2] / "医院器械事件数据.json")
    if not base_path.exists():
        return {"success": False, "error": f"file_not_found: {str(base_path)}"}
    try:
        with base_path.open("r", encoding="utf-8") as fp:
            data_list = _json.load(fp)
    except Exception as e:
        return {"success": False, "error": f"json_read_error: {str(e)}"}
    def norm(s: str) -> str:
        return str(s).strip().lower().replace(" ","")
    zh_map = {
        "医院编号": "hospital_id",
        "医院id": "hospital_id",
        "医院名称": "hospital_id",
        "医疗机构": "hospital_id",
        "器械名称": "device_name",
        "设备名称": "device_name",
        "医疗器械名称": "device_name",
        "制造商": "manufacturer",
        "生产厂家": "manufacturer",
        "型号": "model",
        "批次或序列号": "lot_sn",
        "序列号": "lot_sn",
        "批次": "lot_sn",
        "事件时间": "event_datetime",
        "发现日期": "event_datetime",
        "发现时间": "event_datetime",
        "事件描述": "event_description",
        "事件情况描述": "event_description",
        "不良事件描述": "event_description",
        "伤害严重度": "injury_severity",
        "严重程度": "injury_severity",
        "伤害程度": "injury_severity",
        "处置措施": "action_taken",
        "处置描述": "action_taken",
        "应急处理": "action_taken",
    }
    results = {"received": 0, "duplicate": 0, "failed": 0, "pending": 0, "auto_confirmed": 0}
    ids: List[str] = []
    from datetime import datetime
    for item in (data_list or []):
        try:
            row = {}
            for k, v in item.items():
                key = zh_map.get(k) or zh_map.get(norm(k)) or k
                row[key] = v
            dt = row.get("event_datetime")
            if isinstance(dt, str):
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
                    try:
                        row["event_datetime"] = datetime.strptime(dt, fmt)
                        break
                    except ValueError:
                        continue
            sev = row.get("injury_severity")
            if isinstance(sev, str):
                s = sev.strip().lower()
                cn_map = {"轻微":"mild","轻度":"mild","中度":"moderate","中等":"moderate","重度":"severe","严重":"severe","死亡":"death","无伤害":"none","无":"none"}
                row["injury_severity"] = cn_map.get(s, s)
            if row.get("event_description") is None:
                row["event_description"] = ""
            payload_in = ReportIn(**row)
            out = _process_incoming(payload_in)
            results[out.status] += 1
            ids.append(out.report_id)
            sr = analyze_report_structure(payload_in.dict())
            score = float(sr.get("analysis_confidence", 0) or 0)
            if score >= auto_threshold:
                confirm_structured_report({**payload_in.dict(), "llm_entities": None, "llm_relations": None})
                results["auto_confirmed"] += 1
            elif review_queue:
                from .db import get_db
                get_db().add_pending(out.report_id, {"base": payload_in.dict(), "structure": sr, "score": score})
                results["pending"] += 1
        except Exception as e:
            results["failed"] += 1
            import traceback; traceback.print_exc()
    return {"success": True, "summary": results, "report_ids": ids}
@app.post("/config/neo4j-mapping")
def set_neo4j_mapping(payload: dict = Body(...)):
    import os
    v = bool(payload.get("map_standard_terms", False))
    os.environ["NEO4J_MAP_STANDARD_TERMS"] = "true" if v else "false"
    return {"ok": True, "map_standard_terms": v}

@app.post("/admin/neo4j/clear")
def admin_neo4j_clear():
    out = graph.clear_all()
    if "error" in out:
        raise HTTPException(status_code=503, detail="neo4j_unavailable")
    return out